"""
Parse the APC formulas workbook into structured metric/input/dependency data.

Workbook layout (validated against `APC dashboard Formulas Rev NC.xlsx`):

* Sheet `Inputs`      -> column A holds the canonical input name.
                          Column B optionally holds a sub-classification
                          (Volume, Pick rate, Pack rate, ...).
* Sheet `Formulas <FC>` -> each warehouse has its own sheet.  Column A is the
                          metric name, column B is the primary formula text,
                          and columns C..M hold "layers down" explanations
                          which we keep so users can drill into a metric.

Multiple metrics (Buffer spots CE, VRC, ...) span several rows: the metric
name is set on the first row and subsequent rows have an empty column A.  We
stitch those rows back together so each metric has a single record.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl


@dataclass
class InputItem:
    """A single row from the `Inputs` sheet."""

    name: str           # canonical name (column A)
    qualifier: str = "" # column B (Volume, Pick rate, Actual Pack rate, ...)
    section: str = ""   # last seen section header (CS, VRETs, Transfer out, ...)
    row: int = 0

    @property
    def canonical(self) -> str:
        """Unique identifier combining name + qualifier (qualifier disambiguates
        rows like 'Single Large ' which appears with 'Actual pick rate' and
        also with 'Actual Pack rate')."""

        nm = self.name.strip()
        q = self.qualifier.strip()
        if q and q.lower() not in ("volume",) and q.lower() not in nm.lower():
            return f"{nm} ({q})"
        return nm


@dataclass
class Metric:
    """A single calculated parameter."""

    name: str
    primary_formula: str
    layers: List[str] = field(default_factory=list)
    sheet: str = ""
    row: int = 0

    @property
    def full_text(self) -> str:
        """Concatenated formula + explanations - convenient for token search."""

        parts = [self.primary_formula] + self.layers
        return " | ".join(p for p in parts if p)


@dataclass
class Workbook:
    inputs: List[InputItem]
    metrics_by_warehouse: Dict[str, List[Metric]]
    raw_path: Path

    @property
    def warehouses(self) -> List[str]:
        return sorted(self.metrics_by_warehouse.keys())


SECTION_HEADERS = {
    "CS",
    "VRETs",
    "Transfer out",
    "Transfer in",
    "TSI",
    "TSO",
    "CRETs",
}


def _row_values(ws, row: int, max_col: int) -> List[Optional[object]]:
    return [ws.cell(row, c).value for c in range(1, max_col + 1)]


def _clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_inputs(ws) -> List[InputItem]:
    items: List[InputItem] = []
    current_section = "General"
    for r in range(1, ws.max_row + 1):
        col_a = _clean(ws.cell(r, 1).value)
        col_b = _clean(ws.cell(r, 2).value)

        if not col_a and not col_b:
            continue

        # Detect section headers (single-cell rows like "CS", "VRETs", ...)
        if col_a in SECTION_HEADERS and not col_b:
            current_section = col_a
            continue

        # Skip the column legend row in the file
        if col_a.lower() == "eg" or col_b.lower() == "comments":
            continue
        if col_a == "" and col_b == "":
            continue

        if col_a:
            items.append(
                InputItem(
                    name=col_a,
                    qualifier=col_b,
                    section=current_section,
                    row=r,
                )
            )
    return items


def parse_metrics(ws) -> List[Metric]:
    metrics: List[Metric] = []
    current: Optional[Metric] = None
    max_col = ws.max_column

    for r in range(1, ws.max_row + 1):
        row_vals = _row_values(ws, r, max_col)
        col_a = _clean(row_vals[0])
        col_b = _clean(row_vals[1]) if len(row_vals) > 1 else ""
        rest = [_clean(v) for v in row_vals[2:]] if len(row_vals) > 2 else []

        # Header / legend rows
        if r <= 2:
            continue
        if col_a.startswith("*"):
            continue
        if col_a.lower() == "parameter":
            continue

        if col_a:
            # New metric
            current = Metric(
                name=col_a,
                primary_formula=col_b,
                layers=[v for v in rest if v],
                sheet=ws.title,
                row=r,
            )
            metrics.append(current)
        else:
            # Continuation row: append non-empty content to the most recent
            # metric.  These are typically multi-line formulas (e.g. VRC,
            # Buffer spots CE) where each row contributes another sub-formula.
            if current is None:
                continue
            extras = [v for v in [col_b, *rest] if v]
            if extras:
                current.layers.extend(extras)

    return metrics


def parse_workbook(path: Path) -> Workbook:
    wb = openpyxl.load_workbook(path, data_only=False)
    inputs: List[InputItem] = []
    metrics_by_wh: Dict[str, List[Metric]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name.lower() == "inputs":
            inputs = parse_inputs(ws)
        elif sheet_name.lower().startswith("formulas"):
            # Sheet name pattern: "Formulas RUH8" -> warehouse RUH8.
            wh = sheet_name.split(maxsplit=1)[1] if " " in sheet_name else sheet_name
            metrics_by_wh[wh.strip()] = parse_metrics(ws)

    return Workbook(inputs=inputs, metrics_by_warehouse=metrics_by_wh, raw_path=path)


if __name__ == "__main__":
    import sys

    src = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(
        r"C:\Users\amdtarek\Documents\MCP 2.0\APC dashboard Formulas Rev NC.xlsx"
    )
    book = parse_workbook(src)
    print(f"Inputs: {len(book.inputs)}")
    print(f"Warehouses: {book.warehouses}")
    for wh, metrics in book.metrics_by_warehouse.items():
        print(f"  {wh}: {len(metrics)} metrics")
        for m in metrics[:3]:
            print(f"    - {m.name!r:40s} formula={m.primary_formula[:60]!r}")
